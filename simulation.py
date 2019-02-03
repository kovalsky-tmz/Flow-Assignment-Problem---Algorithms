
import time
import numpy as np
import plotly
from plotly import tools
import plotly.plotly as py
import plotly.graph_objs as go
# USERNAME AND API_KEY TO plot.ly WEBSITE ACCOUNT
# CREATE ACCOUNT ON plot.ly and paste your username and api_key below

# 
from ortools.linear_solver import pywraplp
import pulp
from openpyxl import load_workbook
import random
import math


def linear(start,end,req,cap,cost,flow):

    my_lp_problem = pulp.LpProblem("FA optimizer \nObjective", pulp.LpMinimize)
    for i in range(len(flow)):
        flow[i] = pulp.LpVariable('f{}'.format(i), lowBound=0, cat='Integer')

    # Cel funkcji
    my_lp_problem += sum([flow[i]*cost[i] for i in range(len(flow))]), 'Cost'



    # Ograniczenia "my_lp_problem+="
    my_lp_problem += sum([flow[i] for i in range(len(flow))])>=sum(req)
    for i in range(len(flow)):
        my_lp_problem += flow[i]<=cap[i]


    operation=[]
    for i,(k,v) in enumerate(zip(start,end)):
        tsumplus=[]
        tsumminus=[]

        i_plus=[(j) for j,n in enumerate(start) if n==k]
        for m in i_plus:
            tsumplus.append(flow[m])
            tsumminus.append(req[m])

        i_minus=[(j) for j,n in enumerate(end) if n==k]
        for m in i_minus:
            tsumminus.append(flow[m])
            tsumplus.append(req[m])
        my_lp_problem += (sum(tsumplus))-(sum(tsumminus))==0


    print(my_lp_problem)

    my_lp_problem.solve()

    # WYNIKI
    flow_sum=0
    print(pulp.LpStatus[my_lp_problem.status])
    for variable in my_lp_problem.variables():
        print ("Flow {} amount= {}".format(variable.name, variable.varValue))
        flow_sum+=variable.varValue
    print ('Min Cost : {}'.format(pulp.value(my_lp_problem.objective)))
    delay=((1/sum(req))*((flow_sum)/(sum(cap)-flow_sum)))
    print('Delay : ',round(delay*1000,3),'ms')


###################### TABU ##################################################

def check_cap(cap,flow):
    cond=[]
    for i in range(len(flow)):
        if flow[i]<=cap[i]:
            cond.append(True)
        else:
            cond.append(False)
    if False in cond:
        return False
    else:
        return True 


def check_constraints(start,end,flow,req):
    operation=[]
    cond=[]
    for i,(k,v) in enumerate(zip(start,end)):
        tsumplus=[]
        tsumminus=[]

        i_plus=[(j) for j,n in enumerate(start) if n==k]
        for m in i_plus:
            tsumplus.append(flow[m])
            tsumminus.append(req[m])

        i_minus=[(j) for j,n in enumerate(end) if n==k]
        for m in i_minus:
            tsumminus.append(flow[m])
            tsumplus.append(req[m])
      
        if (sum(tsumplus))-(sum(tsumminus))==0:
            cond.append(True)
        else:
            cond.append(False)

    if False in cond:
        return False
    else:
        return True 



def objective(flow,cost):
    return sum([flow[i]*cost[i] for i in range(len(flow))])

def move_update_tabu(f,t,cap,req):
    i=0
    while  f in t or (not(sum(f)>=sum(req)) or not(check_cap(cap,f)) or not(check_constraints(start,end,f,req))):
        f=[random.randrange(0,cap[i]+1) for i in range(len(cap))]
        # print('nowy: ', f)
        
        if i>40 and len(t)>0:
            t.pop()
        i+=1
    t.append(f)
    
    return f


def tabu(iterations,start,end,req,cap,cost):
    flow=flow_best=[random.randrange(0,cap[i]+1) for i in range(len(cap))]
    bests=[]
    t=[]
    time_tabu=[]
    i=0
    flow_best=move_update_tabu(flow_best,t,cap,req)
    while i<iterations:
        start_time = time.time()
        new=move_update_tabu(flow_best,t,cap,req)
        print('nowy: ', new)
        if(objective(new,cost)<=objective(flow_best,cost)):
            flow_best=new
            print('WYBRANO NAJLEPSZY: ', flow_best)
            print('NAJLEPSZY KOSZT: ', objective(flow_best,cost))
        bests.append(objective(flow_best,cost))    
        i+=1
        time_tabu.append(time.time() - start_time)


    print('optimal flow ',flow_best)
    print('optimal cost: ', objective(flow_best,cost))
    return bests,time_tabu


#####
##### SEARCH LIST################
def move_update(f,cap,req):
    f=[random.randrange(0,cap[i]+1) for i in range(len(cap))]

    while  not(sum(f)>=sum(req)) or not(check_cap(cap,f)) or not(check_constraints(start,end,f,req)):
        f=[random.randrange(0,cap[i]+1) for i in range(len(cap))]
        # print('nowy: ', f)
        
    # print('DODANO: ', f)
    # print('DODANO COST: ', objective(f,cost))
    
    return f

def searching(iterations,start,end,req,cap,cost):
    print("SEARCHING..............")
    flow=flow_best=[random.randrange(0,cap[i]+1) for i in range(len(cap))]
    bests=[]
    time_searching=[]
    i=0
    flow_best=move_update(flow_best,cap,req)
    while i<iterations:
        start_time = time.time()
        new=move_update(flow_best,cap,req)
        # print('nowy: ', new)
        if(objective(new,cost)<=objective(flow_best,cost)):
            flow_best=new
            # print('WYBRANO NAJLEPSZY: ', flow_best)
            print('NAJLEPSZY KOSZT: ', objective(flow_best,cost))
        bests.append(objective(flow_best,cost))    
        i+=1
        time_searching.append(time.time() - start_time)

    print('optimal flow ',flow_best)
    print('optimal cost: ', objective(flow_best,cost))
    return bests,time_searching


# GRAPH FOR METAHEURISTIC ALGORITHMS, HOW IT WORKS INSIDE ALGORITHM
def make_graph_heuristic(i,data_tabu,data_search,time_tabu,time_search):
    
    trace_tabu = go.Scatter(
    x = i,
    y = data_tabu,
    name = 'Cost Tabu',
    line = dict(
        color = ('rgb(205, 12, 24)'),
        width = 4)
    )
    trace_searching = go.Scatter(
        x = i,
        y = data_search,
        name = 'Cost Searching List',
        line = dict(
            color = ('rgb(22, 96, 167)'),
            width = 4,)
        )


    trace_tabu_time = go.Scatter(
    x = i,
    y = time_tabu,
    name = 'Time Tabu',
    line = dict(
        color = ('rgb(205, 12, 24)'),
        width = 4)
    )

    trace_searching_time = go.Scatter(
        x = i,
        y = time_search,
        name = 'Time Searching List',
        line = dict(
            color = ('rgb(22, 96, 167)'),
            width = 4,)
        )

    data = [trace_tabu, trace_searching]
    # Edit the layout
    layout = dict(title = 'Tabu search and List Searching',
                  xaxis = dict(title = 'Iterate'),
                  yaxis = dict(title = 'Cost'),
                  )

    fig = tools.make_subplots(rows=2, cols=2)

    fig.append_trace(trace_tabu, 1, 1)
    fig.append_trace(trace_searching, 1, 2)
    fig.append_trace(trace_tabu_time, 2, 1)
    fig.append_trace(trace_searching_time, 2, 2)

    fig['layout']['xaxis1'].update(title='Iterations')
    fig['layout']['xaxis2'].update(title='Iterations')
    fig['layout']['xaxis3'].update(title='Iterations')
    fig['layout']['xaxis4'].update(title='Iterations')

    fig['layout']['yaxis1'].update(title='Cost')
    fig['layout']['yaxis2'].update(title='Cost')
    fig['layout']['yaxis3'].update(title='Time')
    fig['layout']['yaxis4'].update(title='Time')
    fig['layout'].update(height=1000, width=1000, title='Costs and Times for heuristic algorithms')
    py.iplot(fig, filename='simple-subplot')


#  GRAPH FOR NUMBER OF ITERATIONS AND TIMES
def make_graph(i,l_time,t_time,s_time):
    
    trace_linear = go.Scatter(
    x = i,
    y = l_time,
    name = 'Linear Times',
    line = dict(
        color = ('rgb(205, 12, 24)'),
        width = 4)
    )

    trace_tabu = go.Scatter(
    x = i,
    y = t_time,
    name = 'Tabu Times',
    line = dict(
        color = ('rgb(205, 12, 24)'),
        width = 4)
    )

    trace_searching = go.Scatter(
        x = i,
        y = s_time,
        name = 'Searching List Times',
        line = dict(
            color = ('rgb(22, 96, 167)'),
            width = 4,)
        )

    data = [trace_linear, trace_tabu, trace_searching]
    # Edit the layout
    layout = dict(title = 'Tabu search and List Searching',
                  xaxis = dict(title = 'Iterate'),
                  yaxis = dict(title = 'Time'),
                  )

    fig = tools.make_subplots(rows=2, cols=2)

    fig.append_trace(trace_tabu, 1, 1)
    fig.append_trace(trace_searching, 1, 2)
    fig.append_trace(trace_linear, 2, 1)

    fig['layout']['xaxis1'].update(title='Iterations')
    fig['layout']['xaxis2'].update(title='Iterations')
    fig['layout']['xaxis3'].update(title='Iterations')

    fig['layout']['yaxis1'].update(title='Time')
    fig['layout']['yaxis2'].update(title='Time')
    fig['layout']['yaxis3'].update(title='Time')
    fig['layout'].update(height=1000, width=1000, title='Costs and Times for heuristic algorithms')
    py.iplot(fig, filename='simple-subplot')


def standard_deviation(time):
    average=sum(time)/r
    variance=0

    for i in range(len(time)):
        variance+=(time[i]-average)**2
    variance=variance/len(time)
    standard=math.sqrt(variance)

    return average,standard


########### MAIN APP #####################
if __name__ == '__main__':
	username = input("Enter your username: ")
	api = input("Enter your API KEY: ")

	plotly.tools.set_credentials_file(username=str(username), api_key=str(api))
	start=[]
	end=[]
	req=[]
	cap=[]
	cost=[]
	flow=[]

	r=10 #number of algorithm repeats 
	repeats=[]
	for i in range(1,r+1):
		repeats.append(i)

	k=10 # tabu search and searching list iterations
	iterations=[]
	for i in range(1,k+1):
		iterations.append(i)

	# EXCEL DATA ####################
	wb = load_workbook('./data.xlsx')
	sheet = wb.get_sheet_by_name('Arkusz4') # NAZWA ARKUSZA W EXCELU
	nodes=sheet.cell(row=3, column=2).value

	weights=[[0 for i in range(nodes)] for j in range(nodes)]
	req2=[[0 for i in range(nodes)] for j in range(nodes)]

	for i in range(nodes):
		for j in range(nodes):
			weights[i][j]=(sheet.cell(row=8+i, column=2+j).value)
			if weights[i][j]==1:
				req2[i][j]=(sheet.cell(row=8+i, column=14+j).value)

	for i,v in enumerate(weights):
		for j,v in enumerate(v):
			if v>0:
				start.append(i)
				req.append(req2[i][j])
				end.append(j)

	for i in range(len(start)):
		cap.append(sheet.cell(row=23, column=2+i).value)
		cost.append(sheet.cell(row=25, column=2+i).value)
		flow.append(sheet.cell(row=27, column=2+i).value)

	print('Start Nodes: ',start)
	print('End Nodes: ',end)
	print('Requirements: ',req)
	print('Capacities: ',cap)
	print('Costs: ',cost)
	# EXCEL END ##############################

	# REST
	l_time=[]
	t_time=[]
	s_time=[]

	for i in range(0,r):
		linear_time = time.time()
		linear(start,end,req,cap,cost,flow)
		l_time.append(time.time() - linear_time)
		print("Time: %s seconds" % round((time.time() - linear_time),4))

		tabu_time = time.time()
		tab,time_tabu=tabu(k,start,end,req,cap,cost)
		t_time.append(time.time() - tabu_time)
		print("Time: %s seconds" % round((time.time() - tabu_time),4))

		search_time = time.time()
		search,time_search=searching(k,start,end,req,cap,cost)
		s_time.append(time.time() - search_time)
		print("Time: %s seconds" % round((time.time() - search_time),4))

	# DO WYBORU, JAKI WYKRES GENEROWAĆ
	# make_graph_heuristic(iterations,tab,search,time_tabu,time_search)
	make_graph(repeats,l_time,t_time,s_time)
	# 
	l_average,l_standard=standard_deviation(l_time)
	s_average,s_standard=standard_deviation(s_time)
	t_average,t_standard=standard_deviation(t_time)
	print("Średnia Linear: ",round(l_average,4))
	print("Średnia Search List: ",round(s_average,4))
	print("Średnia Tabu Search: ",round(t_average,4))

	print("Odchylenie standardowe Linear: ",round(l_standard,4))
	print("Odchylenie standardowe Search List: ",round(s_standard,4))
	print("Odchylenie standardowe Tabu Search: ",round(t_standard,4))
